from dotenv import load_dotenv
load_dotenv()  # Cargar variables de entorno

from flask import render_template, request, session, flash, redirect, url_for, jsonify, send_file, Response
from werkzeug.security import generate_password_hash, check_password_hash
from forms.login_form import LoginForm
from flask_login import login_required, current_user, logout_user  # Añadir logout_user
from config import app, db, add_server_timing, cache_get, cache_set
from modelo.models import (
    Usuario,
    Finca,
    Animal,
    Reporte,
    ActividadReciente,
    UsuarioFinca,
    Potrero,
    RotacionPotrero,
    GrupoAnimal,
    Trabajador,
    TipoServicioSalud,
    TipoServicioSexual,
)  # Agregar modelos usados por actualizaciones ORM
from controlador.controlador_actividad import obtener_actividades_recientes  # Importar la función
from datetime import datetime, date  # Añadir esta importación
import time
from sqlalchemy import text, Table, MetaData, update

# Migración segura: asegurar columnas aplica_a_sexo en tablas de tipos de servicio
def ensure_aplica_a_sexo_columns():
    try:
        dialect = db.engine.dialect.name
        if dialect != 'mysql':
            print('Salto migración aplica_a_sexo: dialecto no MySQL ->', dialect)
            return
        with db.engine.connect() as conn:
            # Salud
            exists_salud = conn.execute(text(
                """
                SELECT 1 FROM information_schema.COLUMNS 
                WHERE TABLE_SCHEMA = DATABASE() 
                  AND TABLE_NAME = 'tipo_servicio_salud' 
                  AND COLUMN_NAME = 'aplica_a_sexo'
                """
            )).first()
            if not exists_salud:
                conn.execute(text(
                    """
                    ALTER TABLE tipo_servicio_salud 
                    ADD COLUMN aplica_a_sexo ENUM('macho','hembra','ambos') NULL DEFAULT 'ambos'
                    """
                ))
                print('Columna aplica_a_sexo agregada a tipo_servicio_salud')
        # Inicializar valores nulos vía ORM (seguro y portable)
        try:
            db.session.query(TipoServicioSalud).\
                filter(TipoServicioSalud.aplica_a_sexo == None).\
                update({TipoServicioSalud.aplica_a_sexo: 'ambos'}, synchronize_session=False)
            db.session.commit()
        except Exception:
            db.session.rollback()

        with db.engine.connect() as conn:
            # Sexual
            exists_sexual = conn.execute(text(
                """
                SELECT 1 FROM information_schema.COLUMNS 
                WHERE TABLE_SCHEMA = DATABASE() 
                  AND TABLE_NAME = 'tipo_servicio_sexual' 
                  AND COLUMN_NAME = 'aplica_a_sexo'
                """
            )).first()
            if not exists_sexual:
                conn.execute(text(
                    """
                    ALTER TABLE tipo_servicio_sexual 
                    ADD COLUMN aplica_a_sexo ENUM('macho','hembra','ambos') NULL DEFAULT 'ambos'
                    """
                ))
                print('Columna aplica_a_sexo agregada a tipo_servicio_sexual')
        # Inicializar valores nulos vía ORM
        try:
            db.session.query(TipoServicioSexual).\
                filter(TipoServicioSexual.aplica_a_sexo == None).\
                update({TipoServicioSexual.aplica_a_sexo: 'ambos'}, synchronize_session=False)
            db.session.commit()
        except Exception:
            db.session.rollback()
    except Exception as e:
        # No romper arranque; solo loguear
        print('Error asegurando aplica_a_sexo:', e)

# Migración segura: asegurar columnas de permisos por finca en usuario_finca
def ensure_usuario_finca_permission_columns():
    try:
        dialect = db.engine.dialect.name
        if dialect != 'mysql':
            print('Salto migración usuario_finca: dialecto no MySQL ->', dialect)
            return
        with db.engine.connect() as conn:
            # Verificar columna rol_en_finca
            exists_rol = conn.execute(text(
                """
                SELECT 1 FROM information_schema.COLUMNS 
                WHERE TABLE_SCHEMA = DATABASE() 
                  AND TABLE_NAME = 'usuario_finca' 
                  AND COLUMN_NAME = 'rol_en_finca'
                """
            )).first()
            if not exists_rol:
                conn.execute(text(
                    """
                    ALTER TABLE usuario_finca 
                    ADD COLUMN rol_en_finca SMALLINT NULL
                    """
                ))
                print('Columna rol_en_finca agregada a usuario_finca')

            # Verificar columna puede_editar
            exists_editar = conn.execute(text(
                """
                SELECT 1 FROM information_schema.COLUMNS 
                WHERE TABLE_SCHEMA = DATABASE() 
                  AND TABLE_NAME = 'usuario_finca' 
                  AND COLUMN_NAME = 'puede_editar'
                """
            )).first()
            if not exists_editar:
                conn.execute(text(
                    """
                    ALTER TABLE usuario_finca 
                    ADD COLUMN puede_editar TINYINT(1) NULL DEFAULT 0
                    """
                ))
                print('Columna puede_editar agregada a usuario_finca')

            # Verificar columna estado_asignacion
            exists_estado_asignacion = conn.execute(text(
                """
                SELECT 1 FROM information_schema.COLUMNS 
                WHERE TABLE_SCHEMA = DATABASE() 
                  AND TABLE_NAME = 'usuario_finca' 
                  AND COLUMN_NAME = 'estado_asignacion'
                """
            )).first()
            if not exists_estado_asignacion:
                conn.execute(text(
                    """
                    ALTER TABLE usuario_finca 
                    ADD COLUMN estado_asignacion ENUM('asignado','no_asignado') NOT NULL DEFAULT 'asignado'
                    """
                ))
                print('Columna estado_asignacion agregada a usuario_finca')
        # Inicializar estado_asignacion para filas existentes usando ORM
        try:
            db.session.query(UsuarioFinca).\
                filter(UsuarioFinca.estado_asignacion == None).\
                update({UsuarioFinca.estado_asignacion: 'asignado'}, synchronize_session=False)
            db.session.commit()
        except Exception:
            db.session.rollback()
    except Exception as e:
        print('Error asegurando columnas usuario_finca:', e)

# Migración segura: agregar columna foto_usuario en tabla usuario si no existe
def ensure_foto_usuario_column():
    try:
        dialect = db.engine.dialect.name
        if dialect != 'mysql':
            print('Salto migración foto_usuario: dialecto no MySQL ->', dialect)
            return
        with db.engine.connect() as conn:
            exists_foto = conn.execute(text(
                """
                SELECT 1 FROM information_schema.COLUMNS 
                WHERE TABLE_SCHEMA = DATABASE() 
                  AND TABLE_NAME = 'usuario' 
                  AND COLUMN_NAME = 'foto_usuario'
                """
            )).first()
            if not exists_foto:
                conn.execute(text(
                    """
                    ALTER TABLE usuario 
                    ADD COLUMN foto_usuario LONGBLOB NULL
                    """
                ))
                print('Columna foto_usuario agregada a usuario')
    except Exception as e:
        print('Error asegurando columna foto_usuario:', e)

# Migración segura: crear tabla trabajador si no existe y asegurar índices/unique
def ensure_trabajador_table():
    try:
        dialect = db.engine.dialect.name
        if dialect != 'mysql':
            print('Salto migración trabajador: dialecto no MySQL ->', dialect)
            return
        with db.engine.connect() as conn:
            exists_tbl = conn.execute(text(
                """
                SELECT 1 FROM information_schema.TABLES 
                WHERE TABLE_SCHEMA = DATABASE() 
                  AND TABLE_NAME = 'trabajador'
                """
            )).first()
            if not exists_tbl:
                conn.execute(text(
                    """
                    CREATE TABLE trabajador (
                      id INT AUTO_INCREMENT PRIMARY KEY,
                      usuario_id INT NOT NULL,
                      dueno_id INT NOT NULL,
                      fecha_creacion DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                      estado ENUM('activo','inactivo') NOT NULL DEFAULT 'activo',
                      activo TINYINT(1) NOT NULL DEFAULT 1,
                      CONSTRAINT fk_trabajador_usuario FOREIGN KEY (usuario_id) REFERENCES usuario(id) ON DELETE CASCADE,
                      CONSTRAINT fk_trabajador_dueno FOREIGN KEY (dueno_id) REFERENCES usuario(id) ON DELETE CASCADE,
                      CONSTRAINT uq_trabajador_dueno_usuario UNIQUE (dueno_id, usuario_id)
                    )
                    """
                ))
                print('Tabla trabajador creada')

            # Asegurar columnas clave si la tabla ya existía pero sin ellas
            exists_usuario_col = conn.execute(text(
                """
                SELECT 1 FROM information_schema.COLUMNS 
                WHERE TABLE_SCHEMA = DATABASE() 
                  AND TABLE_NAME = 'trabajador' 
                  AND COLUMN_NAME = 'usuario_id'
                """
            )).first()
            if not exists_usuario_col:
                conn.execute(text("ALTER TABLE trabajador ADD COLUMN usuario_id INT NULL"))
                print('Columna usuario_id agregada a trabajador')

            exists_dueno_col = conn.execute(text(
                """
                SELECT 1 FROM information_schema.COLUMNS 
                WHERE TABLE_SCHEMA = DATABASE() 
                  AND TABLE_NAME = 'trabajador' 
                  AND COLUMN_NAME = 'dueno_id'
                """
            )).first()
            if not exists_dueno_col:
                conn.execute(text("ALTER TABLE trabajador ADD COLUMN dueno_id INT NULL"))
                print('Columna dueno_id agregada a trabajador')

            # Asegurar columna fecha_creacion para evitar errores 1054 al insertar
            exists_fecha_creacion = conn.execute(text(
                """
                SELECT 1 FROM information_schema.COLUMNS 
                WHERE TABLE_SCHEMA = DATABASE() 
                  AND TABLE_NAME = 'trabajador' 
                  AND COLUMN_NAME = 'fecha_creacion'
                """
            )).first()
            if not exists_fecha_creacion:
                conn.execute(text(
                    """
                    ALTER TABLE trabajador 
                    ADD COLUMN fecha_creacion DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
                    """
                ))
                print('Columna fecha_creacion agregada a trabajador')

            # Asegurar columna legacy activo para compatibilidad con el modelo
            exists_activo = conn.execute(text(
                """
                SELECT 1 FROM information_schema.COLUMNS 
                WHERE TABLE_SCHEMA = DATABASE() 
                  AND TABLE_NAME = 'trabajador' 
                  AND COLUMN_NAME = 'activo'
                """
            )).first()
            if not exists_activo:
                conn.execute(text("ALTER TABLE trabajador ADD COLUMN activo TINYINT(1) NOT NULL DEFAULT 1"))
                print('Columna activo agregada a trabajador')

            # Mitigar columna legacy id_jefe (algunos esquemas antiguos la requieren NOT NULL)
            try:
                id_jefe_info = conn.execute(text(
                    """
                    SELECT IS_NULLABLE, COLUMN_DEFAULT 
                    FROM information_schema.COLUMNS 
                    WHERE TABLE_SCHEMA = DATABASE() 
                      AND TABLE_NAME = 'trabajador' 
                      AND COLUMN_NAME = 'id_jefe'
                    """
                )).first()
            except Exception:
                id_jefe_info = None

            if id_jefe_info:
                # Si existe y no permite NULL, flexibilizar para evitar 1364
                try:
                    if id_jefe_info[0] == 'NO':
                        conn.execute(text(
                            """
                            ALTER TABLE trabajador 
                            MODIFY COLUMN id_jefe INT NULL DEFAULT NULL
                            """
                        ))
                        print('Columna legacy id_jefe flexibilizada (NULL DEFAULT NULL)')
                except Exception:
                    pass
                # Opcional: backfill con dueno_id si está NULL para mayor compatibilidad
                try:
                    # Reflejar tabla y actualizar vía SQLAlchemy Core (parametrizado)
                    meta = MetaData()
                    trabajador_tbl = Table('trabajador', meta, autoload_with=conn)
                    stmt = update(trabajador_tbl).\
                        where(trabajador_tbl.c.id_jefe == None).\
                        values({trabajador_tbl.c.id_jefe: trabajador_tbl.c.dueno_id})
                    conn.execute(stmt)
                except Exception:
                    pass

            # Mitigar columna legacy usuario (evitar 1364 si es NOT NULL sin default)
            try:
                usuario_info = conn.execute(text(
                    """
                    SELECT COLUMN_TYPE, IS_NULLABLE, COLUMN_DEFAULT 
                    FROM information_schema.COLUMNS 
                    WHERE TABLE_SCHEMA = DATABASE() 
                      AND TABLE_NAME = 'trabajador' 
                      AND COLUMN_NAME = 'usuario'
                    """
                )).first()
            except Exception:
                usuario_info = None

            if usuario_info:
                col_type, is_nullable, col_default = usuario_info
                try:
                    if is_nullable == 'NO':
                        conn.execute(text(
                            f"""
                            ALTER TABLE trabajador 
                            MODIFY COLUMN usuario {col_type} NULL DEFAULT NULL
                            """
                        ))
                        print('Columna legacy usuario flexibilizada (NULL DEFAULT NULL)')
                except Exception:
                    pass
                # Opcional: normalizar valores vacíos a NULL
                try:
                    meta = MetaData()
                    trabajador_tbl = Table('trabajador', meta, autoload_with=conn)
                    stmt = update(trabajador_tbl).\
                        where(trabajador_tbl.c.usuario == '').\
                        values({trabajador_tbl.c.usuario: None})
                    conn.execute(stmt)
                except Exception:
                    pass

            # Asegurar llaves foráneas si faltan (no falla si ya existen)
            try:
                fk_usuario = conn.execute(text(
                    """
                    SELECT CONSTRAINT_NAME FROM information_schema.KEY_COLUMN_USAGE 
                    WHERE TABLE_SCHEMA = DATABASE() 
                      AND TABLE_NAME = 'trabajador' 
                      AND COLUMN_NAME = 'usuario_id' 
                      AND REFERENCED_TABLE_NAME = 'usuario'
                    """
                )).first()
                if not fk_usuario:
                    conn.execute(text("ALTER TABLE trabajador ADD CONSTRAINT fk_trabajador_usuario FOREIGN KEY (usuario_id) REFERENCES usuario(id) ON DELETE CASCADE"))
            except Exception:
                pass
            try:
                fk_dueno = conn.execute(text(
                    """
                    SELECT CONSTRAINT_NAME FROM information_schema.KEY_COLUMN_USAGE 
                    WHERE TABLE_SCHEMA = DATABASE() 
                      AND TABLE_NAME = 'trabajador' 
                      AND COLUMN_NAME = 'dueno_id' 
                      AND REFERENCED_TABLE_NAME = 'usuario'
                    """
                )).first()
                if not fk_dueno:
                    conn.execute(text("ALTER TABLE trabajador ADD CONSTRAINT fk_trabajador_dueno FOREIGN KEY (dueno_id) REFERENCES usuario(id) ON DELETE CASCADE"))
            except Exception:
                pass

            # Asegurar índices si faltan
            exists_ix_dueno = conn.execute(text(
                """
                SELECT 1 FROM information_schema.STATISTICS 
                WHERE TABLE_SCHEMA = DATABASE() 
                  AND TABLE_NAME = 'trabajador' 
                  AND INDEX_NAME = 'ix_trabajador_dueno_id'
                """
            )).first()
            if not exists_ix_dueno:
                conn.execute(text("CREATE INDEX ix_trabajador_dueno_id ON trabajador(dueno_id)"))
                print('Índice ix_trabajador_dueno_id agregado')

            exists_ix_usuario = conn.execute(text(
                """
                SELECT 1 FROM information_schema.STATISTICS 
                WHERE TABLE_SCHEMA = DATABASE() 
                  AND TABLE_NAME = 'trabajador' 
                  AND INDEX_NAME = 'ix_trabajador_usuario_id'
                """
            )).first()
            if not exists_ix_usuario:
                conn.execute(text("CREATE INDEX ix_trabajador_usuario_id ON trabajador(usuario_id)"))
                print('Índice ix_trabajador_usuario_id agregado')

            # Asegurar unique compuesto
            exists_uq = conn.execute(text(
                """
                SELECT 1 
                FROM information_schema.TABLE_CONSTRAINTS 
                WHERE TABLE_SCHEMA = DATABASE() 
                  AND TABLE_NAME = 'trabajador' 
                  AND CONSTRAINT_NAME = 'uq_trabajador_dueno_usuario' 
                  AND CONSTRAINT_TYPE = 'UNIQUE'
                """
            )).first()
            if not exists_uq:
                conn.execute(text("ALTER TABLE trabajador ADD UNIQUE KEY uq_trabajador_dueno_usuario (dueno_id, usuario_id)"))
                print('Unique uq_trabajador_dueno_usuario agregado')

            # Asegurar columna estado en esquema existente
            exists_estado = conn.execute(text(
                """
                SELECT 1 FROM information_schema.COLUMNS 
                WHERE TABLE_SCHEMA = DATABASE() 
                  AND TABLE_NAME = 'trabajador' 
                  AND COLUMN_NAME = 'estado'
                """
            )).first()
            if not exists_estado:
                conn.execute(text(
                    """
                    ALTER TABLE trabajador 
                    ADD COLUMN estado ENUM('activo','inactivo') NOT NULL DEFAULT 'activo'
                    """
                ))
                print('Columna estado agregada a trabajador')
                # Migrar desde columna legacy activo si existe
                exists_activo_col = conn.execute(text(
                    """
                    SELECT 1 FROM information_schema.COLUMNS 
                    WHERE TABLE_SCHEMA = DATABASE() 
                      AND TABLE_NAME = 'trabajador' 
                      AND COLUMN_NAME = 'activo'
                    """
                )).first()
                if exists_activo_col:
                    meta = MetaData()
                    trabajador_tbl = Table('trabajador', meta, autoload_with=conn)
                    stmt_activo = update(trabajador_tbl).\
                        where(trabajador_tbl.c.activo == 1).\
                        values({trabajador_tbl.c.estado: 'activo'})
                    stmt_inactivo = update(trabajador_tbl).\
                        where(trabajador_tbl.c.activo == 0).\
                        values({trabajador_tbl.c.estado: 'inactivo'})
                    conn.execute(stmt_activo)
                    conn.execute(stmt_inactivo)
    except Exception as e:
        print('Error asegurando tabla trabajador:', e)

# Migración segura: crear tabla permiso_finca_usuario si no existe
def ensure_permiso_finca_usuario_table():
    try:
        dialect = db.engine.dialect.name
        if dialect != 'mysql':
            print('Salto migración permisos_finca: dialecto no MySQL ->', dialect)
            return
        with db.engine.connect() as conn:
            exists_tbl = conn.execute(text(
                """
                SELECT 1 FROM information_schema.TABLES 
                WHERE TABLE_SCHEMA = DATABASE() 
                  AND TABLE_NAME = 'permiso_finca_usuario'
                """
            )).first()
            if not exists_tbl:
                conn.execute(text(
                    """
                    CREATE TABLE permiso_finca_usuario (
                      id INT AUTO_INCREMENT PRIMARY KEY,
                      trabajador_id INT NOT NULL,
                      finca_id SMALLINT NOT NULL,
                      crear_potreros TINYINT(1) NOT NULL DEFAULT 0,
                      agregar_animales TINYINT(1) NOT NULL DEFAULT 0,
                      eliminar_animales TINYINT(1) NOT NULL DEFAULT 0,
                      crear_usuarios_ligados TINYINT(1) NOT NULL DEFAULT 0,
                      actualizar_datos_usuario TINYINT(1) NOT NULL DEFAULT 0,
                      updated_at DATETIME NULL,
                      CONSTRAINT fk_perm_trabajador FOREIGN KEY (trabajador_id) REFERENCES trabajador(id_trabajador) ON DELETE CASCADE,
                      CONSTRAINT fk_perm_finca FOREIGN KEY (finca_id) REFERENCES finca(id_finca) ON DELETE CASCADE,
                      CONSTRAINT uq_permiso_finca_trabajador UNIQUE (trabajador_id, finca_id)
                    )
                    """
                ))
                print('Tabla permiso_finca_usuario creada (FK trabajador)')
            else:
                # Si existe la tabla, asegurar columna trabajador_id y constraints/índices
                has_trab_col = conn.execute(text(
                    """
                    SELECT 1 FROM information_schema.COLUMNS
                    WHERE TABLE_SCHEMA = DATABASE()
                      AND TABLE_NAME = 'permiso_finca_usuario'
                      AND COLUMN_NAME = 'trabajador_id'
                    """
                )).first()
                if not has_trab_col:
                    try:
                        conn.execute(text("ALTER TABLE permiso_finca_usuario ADD COLUMN trabajador_id INT NULL"))
                        print('Columna trabajador_id agregada a permiso_finca_usuario')
                    except Exception:
                        pass
                    # Agregar FK si falta
                    try:
                        conn.execute(text("ALTER TABLE permiso_finca_usuario ADD CONSTRAINT fk_perm_trabajador FOREIGN KEY (trabajador_id) REFERENCES trabajador(id_trabajador) ON DELETE CASCADE"))
                    except Exception:
                        pass
                    # Agregar índice sobre trabajador_id
                    try:
                        conn.execute(text("CREATE INDEX ix_permiso_trabajador ON permiso_finca_usuario(trabajador_id)"))
                    except Exception:
                        pass
                    # Agregar unique compuesto (no falla si ya existe otro)
                    try:
                        conn.execute(text("ALTER TABLE permiso_finca_usuario ADD UNIQUE KEY uq_permiso_finca_trabajador (trabajador_id, finca_id)"))
                    except Exception:
                        pass
            # Asegurar índice de finca
            try:
                conn.execute(text("CREATE INDEX ix_permiso_finca ON permiso_finca_usuario(finca_id)"))
            except Exception:
                pass
    except Exception as e:
        print('Error asegurando tabla permiso_finca_usuario:', e)

# Migración segura: asegurar columna fecha_proximo en servicios_sexuales
from sqlalchemy import text

def ensure_fecha_proximo_servicios_sexuales():
    try:
        dialect = db.engine.dialect.name
        if dialect != 'mysql':
            print('Salto migración fecha_proximo (servicios_sexuales): dialecto no MySQL ->', dialect)
            return
        with db.engine.connect() as conn:
            exists_col = conn.execute(text(
                """
                SELECT 1 FROM information_schema.COLUMNS 
                WHERE TABLE_SCHEMA = DATABASE() 
                  AND TABLE_NAME = 'servicios_sexuales' 
                  AND COLUMN_NAME = 'fecha_proximo'
                """
            )).first()
            if not exists_col:
                conn.execute(text(
                    """
                    ALTER TABLE servicios_sexuales 
                    ADD COLUMN fecha_proximo DATE NULL
                    """
                ))
                print('Columna fecha_proximo agregada a servicios_sexuales')
    except Exception as e:
        print('Error asegurando fecha_proximo en servicios_sexuales:', e)

# Crear todas las tablas en la base de datos
with app.app_context():
    db.create_all()
    print("Tablas creadas correctamente en la base de datos")
    # Ejecutar migración segura de columnas aplica_a_sexo
    ensure_aplica_a_sexo_columns()
    # Ejecutar migración segura de columnas de usuario_finca
    ensure_usuario_finca_permission_columns()
    # Ejecutar migración segura de foto_usuario
    ensure_foto_usuario_column()
    # Ejecutar migración segura de tabla trabajador (añade columna 'estado' si falta)
    ensure_trabajador_table()
    ensure_permiso_finca_usuario_table()
    # Nueva: asegurar columna fecha_proximo en servicios_sexuales
    ensure_fecha_proximo_servicios_sexuales()
    
    # Importar e inicializar el usuario administrador
    from modelo.models import inicializar_usuario_admin
    inicializar_usuario_admin()

# Importar las funciones después de definir las rutas para evitar la importación circular
@app.route('/')
def pagina_inicio():
    return render_template('pages/inicio.html')

# Definir las rutas primero
@app.route('/dashboard/root')
def dashboard_root():
    # Obtener estadísticas desde la base de datos
    total_usuarios = Usuario.query.count()
    total_fincas = Finca.query.count()
    total_ganado = Animal.query.count()
    total_reportes = Reporte.query.count()
    
    # Obtener actividades recientes
    actividades_recientes = obtener_actividades_recientes(5)  # Obtener las 5 actividades más recientes
    
    return render_template('root/dashboard_root.html', 
                          total_usuarios=total_usuarios,
                          total_fincas=total_fincas,
                          total_ganado=total_ganado,
                          total_reportes=total_reportes,
                          actividades_recientes=actividades_recientes)  # Pasar actividades a la plantilla

# En la ruta del dashboard:
@app.route('/dashboard/dueno')
@login_required
def dashboard_dueno():
    # Obtener el usuario actual
    usuario_actual = Usuario.query.get(current_user.id)
    
    # Cache de conteos por usuario para reducir TTFB
    counts_key = f"counts_user:{current_user.id}"
    _tc = time.perf_counter()
    cached_counts = cache_get(counts_key)
    if cached_counts:
        total_fincas = cached_counts.get('total_fincas', 0)
        total_animales = cached_counts.get('total_animales', 0)
        total_trabajadores = cached_counts.get('total_trabajadores', 0)
        add_server_timing('cache_counts_hit', (time.perf_counter() - _tc) * 1000.0)
    else:
        add_server_timing('cache_counts_miss', (time.perf_counter() - _tc) * 1000.0)
        from sqlalchemy import func
        # Optimizado: usar subconsulta de las fincas del dueño con DISTINCT y COUNT
        _t0 = time.perf_counter()
        fincas_dueno_subq = db.session.query(UsuarioFinca.finca_id).\
            filter(UsuarioFinca.usuario_id == current_user.id).distinct()
        total_fincas = db.session.query(func.count()).select_from(fincas_dueno_subq).scalar() or 0
        add_server_timing('q_fincas_count', (time.perf_counter() - _t0) * 1000.0)
        
        # Contar los animales en las fincas del usuario sin seleccionar todas las columnas
        _t1 = time.perf_counter()
        total_animales = db.session.query(func.count(Animal.id_animal)).\
            filter(Animal.id_finca.in_(fincas_dueno_subq)).scalar() or 0
        add_server_timing('q_animales_count', (time.perf_counter() - _t1) * 1000.0)
        
        # Contar trabajadores: preferir Trabajador con estado 'activo', fallback seguro a UsuarioFinca
        try:
            _t2 = time.perf_counter()
            total_trabajadores = db.session.query(func.count(Trabajador.id_trabajador)).\
                filter(Trabajador.id_jefe == current_user.id, Trabajador.estado == 'activo').scalar() or 0
            add_server_timing('q_trabajadores_count', (time.perf_counter() - _t2) * 1000.0)
        except Exception:
            # Fallback: contar usuarios relacionados a fincas del dueño que no sean el dueño mismo
            _t2b = time.perf_counter()
            total_trabajadores = db.session.query(func.count(UsuarioFinca.usuario_id)).\
                filter(
                    UsuarioFinca.finca_id.in_(fincas_dueno_subq),
                    UsuarioFinca.usuario_id != current_user.id
                ).scalar() or 0
            add_server_timing('q_relaciones_count', (time.perf_counter() - _t2b) * 1000.0)

        # Guardar conteos en caché por 60s
        cache_set(counts_key, {
            'total_fincas': total_fincas,
            'total_animales': total_animales,
            'total_trabajadores': total_trabajadores,
        }, timeout=60)

    # Definir total_produccion (ajusta esto según tu modelo de datos)
    total_produccion = 0  # Inicializar con un valor predeterminado o calcular según tus necesidades

    # Obtener actividades recientes del usuario (cache 30s)
    acts_key = f"acts_user:{current_user.id}"
    _t3 = time.perf_counter()
    actividades_recientes = cache_get(acts_key)
    if actividades_recientes is None:
        actividades_recientes = ActividadReciente.query.filter_by(usuario_id=current_user.id).order_by(ActividadReciente.fecha.desc()).limit(5).all()
        cache_set(acts_key, actividades_recientes, timeout=30)
        add_server_timing('q_actividades_recientes', (time.perf_counter() - _t3) * 1000.0)
    else:
        add_server_timing('cache_acts_hit', (time.perf_counter() - _t3) * 1000.0)
    
    # Añadir la fecha y hora actual
    now = datetime.now()
    
    _t_render = time.perf_counter()
    html = render_template('dueño/dashboard_dueno.html', 
                           total_fincas=total_fincas,
                           total_animales=total_animales,
                           total_produccion=total_produccion,
                           total_trabajadores=total_trabajadores,
                           now=now)
    add_server_timing('render_template', (time.perf_counter() - _t_render) * 1000.0)
    return html

@app.route('/dashboard/trabajador')
def dashboard_trabajador():
    # Obtener métricas visibles para el trabajador basadas en sus fincas asignadas
    from sqlalchemy import func
    usuario_actual = Usuario.query.get(current_user.id) if 'usuario_id' in session else None

    # Subconsulta de fincas asignadas al trabajador
    fincas_asignadas_subq = db.session.query(UsuarioFinca.finca_id).\
        filter(UsuarioFinca.usuario_id == current_user.id).distinct()

    total_fincas = db.session.query(func.count()).select_from(fincas_asignadas_subq).scalar() or 0
    total_animales = db.session.query(func.count(Animal.id_animal)).\
        filter(Animal.id_finca.in_(fincas_asignadas_subq)).scalar() or 0
    total_produccion = 0
    # Placeholder: contar tareas si existe modelo, si no, usar 0
    total_tareas = 0

    # Actividad reciente del trabajador
    actividades_recientes = ActividadReciente.query.filter_by(usuario_id=current_user.id).\
        order_by(ActividadReciente.fecha.desc()).limit(5).all()

    from datetime import datetime
    now = datetime.now()

    return render_template('trabajador-veternario/dashboard_trabajador.html',
                           total_fincas=total_fincas,
                           total_animales=total_animales,
                           total_produccion=total_produccion,
                           total_tareas=total_tareas,
                           actividades_recientes=actividades_recientes,
                           now=now)

# Ahora importar las funciones de autenticación
# Agregar después de la importación de controlador_autenticacion
from controlador.controlador_autenticacion import ruta_login, ruta_logout, requiere_rol, ruta_registro, configurar_google_oauth, google_login, cambiar_password

# Configurar Google OAuth
google_blueprint = configurar_google_oauth(app)

# Agregar la ruta de registro
@app.route('/registro', methods=['GET', 'POST'])
def registro():
    return ruta_registro()

# Y aplicar los decoradores después de importar
@app.route('/login', methods=['GET', 'POST'])
def login():
    return ruta_login()

@app.route('/logout')
def logout():
    return ruta_logout()

# Añadir ruta para el login con Google
@app.route('/login/google')
def login_google():
    return google_login()

# Ruta de cambio de contraseña obligatoria
@app.route('/cambiar-password', methods=['GET', 'POST'])
def cambiar_password_route():
    return cambiar_password()

# Importar controlador de usuarios
from controlador.controlador_usuario import listar_usuarios, crear_usuario, editar_usuario, eliminar_usuario_controlador

# Importar las funciones de gestión de animales
from controlador.controlador_animal import (
    listar_animales, crear_animal, editar_animal, eliminar_animal, ver_animal,
    obtener_animales_por_finca, obtener_animales_por_sexo, get_potreros_por_finca,
    get_animales_disponibles, get_animales_potrero, asignar_animales_potrero,
    ver_foto_animal, ver_animales_finca, ver_animales_fuera, ver_animales_fuera_global,
    api_madurez_sexual_por_raza, documentos_geneticos, agregar_documento_genetico,
    ver_documento_genetico, descargar_documento_genetico, eliminar_documento_genetico,
    procedimientos_animal, eliminar_servicio_salud, eliminar_servicio_sexual, eliminar_registro_peso,
    genealogia_animal, historial_procedimientos, gestion_produccion, gestion_produccion_finca,
    consumo_animal, biologicos_animal, ver_produccion_animal,
    # Importar vistas dedicadas
    ver_peso_animal, ver_ciclo_animal, ver_salud_animal, ver_graficos_animal
)
from controlador.controlador_animal import editar_documento_genetico

# Importar las funciones de gestión de fincas
from controlador.controlador_finca import (
    crear_finca, editar_finca, eliminar_finca, listar_fincas, gestionar_finca,
    obtener_fincas_usuario, ver_finca, crear_potrero, editar_potrero,
    eliminar_potrero, ver_potrero, agregar_animales_potrero,
    listar_grupos_finca, gestionar_grupo, api_agregar_animal_a_grupo, api_quitar_animal_de_grupo, eliminar_grupo
)
from controlador.controlador_finca import guardar_rotacion
from controlador.controlador_finca import api_default_tipo_uso_potrero

# Importar controlador de compras
from controlador.controlador_compra import *
from controlador.controlador_trabajador import listar_trabajadores_dueno, crear_trabajador_dueno, actualizar_permiso_trabajador, asignar_trabajador_finca, quitar_trabajador_finca, obtener_permisos_finca_trabajador, actualizar_permiso_finca_trabajador, gestionar_trabajadores_finca
from controlador.controlador_trabajador_admin import ver_trabajador_admin, actualizar_trabajador_admin, ver_foto_usuario, eliminar_trabajador

# Aplicar los decoradores de rol a las rutas ya definidas
dashboard_root = requiere_rol(3)(dashboard_root)  # Solo accesible para rol 3 (root)
dashboard_dueno = requiere_rol(2)(dashboard_dueno)  # Accesible para roles 2 y 3
dashboard_trabajador = requiere_rol(1)(dashboard_trabajador)  # Accesible para roles 1, 2 y 3
dashboard_trabajador = requiere_rol(1)(dashboard_trabajador)  # Accesible para roles 1, 2 y 3

# Asegurar que los decoradores aplicados después de la definición afecten las rutas registradas
app.view_functions['dashboard_root'] = login_required(dashboard_root)
app.view_functions['dashboard_dueno'] = dashboard_dueno
app.view_functions['dashboard_trabajador'] = login_required(dashboard_trabajador)

# Rutas de gestión de usuarios (solo para admin)
@app.route('/admin/usuarios')
@login_required
def admin_usuarios():
    return listar_usuarios()

# Vista de reportes para administrador root
@app.route('/admin/reportes')
@login_required
def admin_reportes():
    # Datos base para filtros (fincas del sistema)
    fincas = Finca.query.all()
    reportes_def = [
        { 'id': 'resumen', 'nombre': 'Resumen del sistema' },
        { 'id': 'usuarios_rol', 'nombre': 'Usuarios por rol' },
        { 'id': 'fincas_dueno', 'nombre': 'Fincas por dueño' },
        { 'id': 'animales_finca', 'nombre': 'Animales por finca' },
        { 'id': 'animales_sexo', 'nombre': 'Animales por sexo' },
        { 'id': 'produccion_finca', 'nombre': 'Producción por finca' },
    ]
    return render_template('root/reportes.html', fincas=fincas, reportes_def=reportes_def)

# Vista de reportes para dueño de finca
@app.route('/reportes', endpoint='reportes_dueno')
@login_required
@requiere_rol(2)
def reportes_dueno():
    # Fincas del dueño actual
    from modelo.models import Finca, UsuarioFinca
    fincas = Finca.query.join(UsuarioFinca).filter(UsuarioFinca.usuario_id == current_user.id).all()

    # Definición inicial de tipos de reporte para el dueño
    reportes_def = [
        { 'id': 'resumen_personal', 'nombre': 'Resumen de mis fincas' },
        { 'id': 'animales_finca', 'nombre': 'Animales por finca' },
        { 'id': 'animales_sexo', 'nombre': 'Animales por sexo' },
        { 'id': 'produccion_finca', 'nombre': 'Producción por finca' },
        { 'id': 'procedimientos_por_tipo', 'nombre': 'Procedimientos por tipo' },
    ]

    return render_template('dueño/reportes.html', fincas=fincas, reportes_def=reportes_def)

# Utilidades para exportación de reportes del dueño
def _parse_date(val):
    if not val:
        return None
    try:
        # Espera formato yyyy-mm-dd del input type="date"
        return datetime.strptime(val, '%Y-%m-%d').date()
    except Exception:
        return None

def _dataset_dueno(reporte_tipo: str, finca_id: int | None, desde: date | None, hasta: date | None):
    from modelo.models import Animal, ProductosAnimal, Productos, Finca, UsuarioFinca, Raza, Potrero
    # Asegurar que solo se consulten fincas del dueño
    finca_ids = [rel.finca_id for rel in UsuarioFinca.query.filter_by(usuario_id=current_user.id).all()]
    def within_owner(query):
        return query.filter(Animal.id_finca.in_(finca_ids))

    if reporte_tipo == 'animales_finca':
        q = Animal.query
        q = within_owner(q)
        if finca_id and finca_id != 0:
            q = q.filter(Animal.id_finca == finca_id)
        animales = q.all()
        data = []
        for a in animales:
            # Raza puede ser None si datos inconsistentes
            raza_nombre = getattr(a.raza, 'nombre_raza', '') if hasattr(a, 'raza') else ''
            data.append({
                'ID': a.id_animal,
                'Nombre': a.nombre_animal,
                'Sexo': a.sexo,
                'Raza': raza_nombre,
                'Finca': getattr(a.finca, 'nombre_finca', ''),
                'Ubicación': a.ubicacion_animal
            })
        return {
            'titulo': 'Animales por finca',
            'columns': ['ID', 'Nombre', 'Sexo', 'Raza', 'Finca', 'Ubicación'],
            'rows': data,
        }
    elif reporte_tipo == 'produccion_finca':
        # Detalle de producción con joins
        q = ProductosAnimal.query.join(Animal, ProductosAnimal.id_animal == Animal.id_animal).join(Productos, ProductosAnimal.id_producto == Productos.id_producto)
        q = q.filter(Animal.id_finca.in_(finca_ids))
        if finca_id and finca_id != 0:
            q = q.filter(Animal.id_finca == finca_id)
        if desde:
            q = q.filter(ProductosAnimal.fecha >= desde)
        if hasta:
            q = q.filter(ProductosAnimal.fecha <= hasta)
        regs = q.all()
        data = []
        total = 0.0
        for r in regs:
            finca_nombre = getattr(r.animal.finca, 'nombre_finca', '') if hasattr(r.animal, 'finca') else ''
            producto = getattr(r.producto, 'nombre_producto', '') if hasattr(r, 'producto') else ''
            total += float(getattr(r, 'cantidad', 0) or 0)
            data.append({
                'Fecha': r.fecha.isoformat(),
                'Producto': producto,
                'Cantidad': float(getattr(r, 'cantidad', 0) or 0),
                'Animal': getattr(r.animal, 'nombre_animal', ''),
                'Finca': finca_nombre,
            })
        return {
            'titulo': 'Producción por finca',
            'columns': ['Fecha', 'Producto', 'Cantidad', 'Animal', 'Finca'],
            'rows': data,
            'meta': { 'total_cantidad': total }
        }
    elif reporte_tipo == 'resumen_personal':
        # Resumen enriquecido por finca del dueño
        fincas = Finca.query.filter(Finca.id_finca.in_(finca_ids)).all()
        rows = []
        # Acumuladores globales para totales y promedios
        totals = {
            'Potreros': 0,
            'Animales': 0,
            'Machos': 0,
            'Hembras': 0,
            'En finca': 0,
            'Fuera de finca': 0,
            'Sin potrero': 0,
        }
        total_age_years_sum = 0.0
        total_age_count = 0
        global_razas_set = set()
        global_produce_set = set()

        for f in fincas:
            # Cargar animales de la finca para cálculos derivados
            animales_finca = Animal.query.filter(Animal.id_finca == f.id_finca).all()
            total_animales = len(animales_finca)
            machos = sum(1 for a in animales_finca if (a.sexo or '').lower() == 'macho')
            hembras = sum(1 for a in animales_finca if (a.sexo or '').lower() == 'hembra')
            en_finca = sum(1 for a in animales_finca if (a.ubicacion_animal or '').lower() == 'en finca')
            fuera_finca = sum(1 for a in animales_finca if (a.ubicacion_animal or '').lower() == 'fuera de la finca')
            sin_potrero = sum(1 for a in animales_finca if not getattr(a, 'id_potrero', None))

            # Edad media (años)
            edades = []
            for a in animales_finca:
                if getattr(a, 'fecha_nacimiento', None):
                    try:
                        edades.append((date.today() - a.fecha_nacimiento).days / 365.25)
                    except Exception:
                        pass
                # Razas distintas a nivel global
                if getattr(a, 'id_raza', None):
                    global_razas_set.add(a.id_raza)
            edad_media = round(sum(edades) / len(edades), 2) if edades else 0.0

            # Potreros en la finca
            potreros = Potrero.query.filter(Potrero.id_finca == f.id_finca).count()

            # Tipos de producción (texto): leche, carne, estiercol
            try:
                q_prod = (
                    db.session.query(Productos.nombre_producto)
                    .join(ProductosAnimal, Productos.id_producto == ProductosAnimal.id_producto)
                    .join(Animal, ProductosAnimal.id_animal == Animal.id_animal)
                    .filter(Animal.id_finca == f.id_finca)
                )
                if desde:
                    q_prod = q_prod.filter(ProductosAnimal.fecha >= desde)
                if hasta:
                    q_prod = q_prod.filter(ProductosAnimal.fecha <= hasta)
                nombres_prod = [n[0] for n in q_prod.distinct().all()]
            except Exception:
                nombres_prod = []

            def _cat(nombre: str) -> str | None:
                n = (nombre or '').strip().lower()
                if 'leche' in n:
                    return 'leche'
                if 'carne' in n:
                    return 'carne'
                # cubrir "estiércol" y "estiercol"
                if 'estiércol' in n or 'estiercol' in n:
                    return 'estiercol'
                # animal vivo
                if 'animal vivo' in n or (n == 'animal'):
                    return 'animal vivo'
                # biológicos registrados mediante ProductosAnimal
                if 'semen' in n:
                    return 'semen'
                if 'embrion' in n or 'embri' in n:
                    return 'embrion'
                return None

            cats = {c for c in (_cat(x) for x in nombres_prod) if c}
            global_produce_set.update(cats)
            produce_text = ', '.join(sorted(cats)) if cats else '—'

            # Registrar fila
            rows.append({
                'Finca': f.nombre_finca,
                'Potreros': potreros,
                'Animales': total_animales,
                'Machos': machos,
                'Hembras': hembras,
                'En finca': en_finca,
                'Fuera de finca': fuera_finca,
                'Sin potrero': sin_potrero,
                'Edad media (años)': edad_media,
                'Razas': len({a.id_raza for a in animales_finca if getattr(a, 'id_raza', None)}),
                'Se produce': produce_text,
            })

            # Actualizar acumuladores globales
            totals['Potreros'] += potreros
            totals['Animales'] += total_animales
            totals['Machos'] += machos
            totals['Hembras'] += hembras
            totals['En finca'] += en_finca
            totals['Fuera de finca'] += fuera_finca
            totals['Sin potrero'] += sin_potrero
            total_age_years_sum += sum(edades)
            total_age_count += len(edades)

        # Totales y promedios globales
        totals['Edad media (años)'] = round(total_age_years_sum / total_age_count, 2) if total_age_count else 0.0
        totals['Razas'] = len(global_razas_set)
        totals['Se produce'] = ', '.join(sorted(global_produce_set)) if global_produce_set else '—'

        return {
            'titulo': 'Resumen de mis fincas',
            'columns': [
                'Finca', 'Potreros', 'Animales', 'Machos', 'Hembras',
                'En finca', 'Fuera de finca', 'Sin potrero',
                'Edad media (años)', 'Razas', 'Se produce'
            ],
            'rows': rows,
            'meta': {
                'totals': {
                    'Finca': 'TOTAL',
                    **totals
                }
            }
        }
    else:
        # Fallback: devolver estructura vacía
        return {
            'titulo': 'Reporte',
            'columns': [],
            'rows': []
        }

def _make_excel_response(dataset: dict, filename: str):
    # Preferir openpyxl para Excel estilizado; fallback a XlsxWriter y luego CSV
    import io
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = 'Datos'

        # Estilos
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
        thin = Side(style='thin', color='999999')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Título y fecha
        ws.cell(row=1, column=1, value=dataset.get('titulo', 'Reporte')).font = title_font
        ws.cell(row=2, column=1, value=f"Generado: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}")

        cols = dataset.get('columns', [])
        rows = dataset.get('rows', [])

        # Encabezados
        for c_idx, col in enumerate(cols, start=1):
            cell = ws.cell(row=3, column=c_idx, value=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='left')

        # Datos
        for r_idx, row in enumerate(rows, start=4):
            for c_idx, col in enumerate(cols, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=row.get(col, ''))
                cell.border = border
                cell.alignment = Alignment(horizontal='left')

        last_row = 3 + max(1, len(rows) + 1)

        # Totales si existen
        meta = dataset.get('meta') or {}
        totals = meta.get('totals') or {}
        if totals and cols:
            total_row = 3 + len(rows) + 1
            for c_idx, col in enumerate(cols, start=1):
                cell = ws.cell(row=total_row, column=c_idx, value=totals.get(col, ''))
                cell.font = Font(bold=True)
                cell.border = border

        # Freeze panes y filtro
        ws.freeze_panes = 'A4'
        if cols:
            last_col_letter = get_column_letter(len(cols))
            ws.auto_filter.ref = f"A3:{last_col_letter}{3 + len(rows)}"

        # Auto ancho según contenido
        for c_idx, col in enumerate(cols, start=1):
            max_len = len(str(col))
            for r in rows:
                max_len = max(max_len, len(str(r.get(col, ''))))
            ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 2, 30)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name=f"{filename}.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception:
        try:
            import xlsxwriter
            output = io.BytesIO()
            workbook = xlsxwriter.Workbook(output, {'in_memory': True})
            ws = workbook.add_worksheet('Datos')
            title_fmt = workbook.add_format({'bold': True, 'font_size': 14})
            header_fmt = workbook.add_format({'bold': True, 'bg_color': '#D9EAD3', 'border': 1})
            cell_fmt = workbook.add_format({'border': 1})
            total_fmt = workbook.add_format({'bold': True, 'border': 1})
            # Título
            ws.write(0, 0, dataset.get('titulo', 'Reporte'), title_fmt)
            # Fecha de generación
            ws.write(1, 0, f"Generado: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}")
            # Encabezados
            for col_idx, col in enumerate(dataset.get('columns', [])):
                ws.write(2, col_idx, col, header_fmt)
            # Filas
            for row_idx, row in enumerate(dataset.get('rows', []), start=3):
                for col_idx, col in enumerate(dataset.get('columns', [])):
                    ws.write(row_idx, col_idx, row.get(col, ''), cell_fmt)
            last_data_row = 3 + len(dataset.get('rows', []))
            # Auto ancho
            for col_idx in range(len(dataset.get('columns', []))):
                ws.set_column(col_idx, col_idx, 18)
            # Filtro y congelar encabezado
            ws.autofilter(2, 0, max(2, last_data_row), max(0, len(dataset.get('columns', [])) - 1))
            ws.freeze_panes(3, 0)
            # Resumen si aplica
            meta = dataset.get('meta') or {}
            if 'total_cantidad' in meta:
                ws.write(1, 0, 'Total cantidad', header_fmt)
                ws.write(1, 1, meta['total_cantidad'], cell_fmt)
            # Fila de totales organizada según columnas
            totals = (meta.get('totals') or {})
            if totals and dataset.get('columns'):
                for col_idx, col in enumerate(dataset.get('columns', [])):
                    ws.write(last_data_row, col_idx, totals.get(col, ''), total_fmt)
            workbook.close()
            output.seek(0)
            return send_file(output, as_attachment=True, download_name=f"{filename}.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        except Exception:
            # Fallback CSV
            import csv
            si = io.StringIO()
            writer = csv.writer(si)
            writer.writerow([dataset.get('titulo', 'Reporte')])
            writer.writerow([f"Generado: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"])
            writer.writerow(dataset.get('columns', []))
            for row in dataset.get('rows', []):
                writer.writerow([row.get(col, '') for col in dataset.get('columns', [])])
            totals = (dataset.get('meta', {}) or {}).get('totals')
            if totals and dataset.get('columns'):
                writer.writerow([totals.get(col, '') for col in dataset.get('columns', [])])
            output = io.BytesIO(si.getvalue().encode('utf-8'))
            return send_file(output, as_attachment=True, download_name=f"{filename}.csv", mimetype='text/csv')

def _make_pdf_response(dataset: dict, filename: str):
    # Intentar con reportlab para PDF estilizado; si no está, usar HTML como fallback
    try:
        import io
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        title = Paragraph(dataset.get('titulo', 'Reporte'), styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 12))
        cols = dataset.get('columns', [])
        rows = dataset.get('rows', [])
        table_data = [cols] + [[row.get(c, '') for c in cols] for row in rows]
        tbl = Table(table_data)
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#D9EAD3')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.black),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ]))
        elements.append(tbl)
        doc.build(elements)
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name=f"{filename}.pdf", mimetype='application/pdf')
    except Exception:
        # Fallback sencillo a HTML
        html = ['<html><head><meta charset="utf-8"><title>Reporte</title><style>table{border-collapse:collapse;width:100%}th,td{border:1px solid #ccc;padding:8px}th{background:#D9EAD3}</style></head><body>']
        html.append(f"<h2>{dataset.get('titulo','Reporte')}</h2>")
        cols = dataset.get('columns', [])
        html.append('<table>')
        if cols:
            html.append('<thead><tr>' + ''.join(f'<th>{c}</th>' for c in cols) + '</tr></thead>')
        html.append('<tbody>')
        for row in dataset.get('rows', []):
            html.append('<tr>' + ''.join(f'<td>{row.get(c, '')}</td>' for c in cols) + '</tr>')
        html.append('</tbody></table></body></html>')
        return Response('\n'.join(html), mimetype='text/html')

# Dataset JSON para vista previa (dueño)
@app.route('/reportes/dataset', methods=['GET'], endpoint='dataset_reporte_dueno')
@login_required
@requiere_rol(2)
def dataset_reporte_dueno():
    """Devuelve el dataset del reporte en formato JSON para vista previa en línea."""
    reporte_tipo = request.args.get('reporte_tipo', 'animales_finca')
    finca_id = request.args.get('finca_id', type=int)
    desde = _parse_date(request.args.get('desde'))
    hasta = _parse_date(request.args.get('hasta'))
    dataset = _dataset_dueno(reporte_tipo, finca_id, desde, hasta)
    dataset['generated_at'] = datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')
    return jsonify(dataset)

# Exportación Excel para dueño
@app.route('/reportes/export/excel', methods=['GET'], endpoint='exportar_reporte_excel_dueno')
@login_required
@requiere_rol(2)
def exportar_reporte_excel_dueno():
    reporte_tipo = request.args.get('reporte_tipo', 'animales_finca')
    finca_id = request.args.get('finca_id', type=int)
    desde = _parse_date(request.args.get('desde'))
    hasta = _parse_date(request.args.get('hasta'))
    dataset = _dataset_dueno(reporte_tipo, finca_id, desde, hasta)
    # Solicitud: remover "Producción total" únicamente del Excel en "Resumen de mis fincas"
    if reporte_tipo == 'resumen_personal':
        col_remove = 'Producción total'
        # Filtrar columnas
        dataset['columns'] = [c for c in (dataset.get('columns') or []) if c != col_remove]
        # Remover campo de cada fila
        for r in (dataset.get('rows') or []):
            if isinstance(r, dict):
                r.pop(col_remove, None)
        # Remover de totales si existe
        meta = dataset.get('meta') or {}
        totals = meta.get('totals')
        if isinstance(totals, dict):
            totals.pop(col_remove, None)
        dataset['meta'] = meta
    filename = f"{reporte_tipo}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}"
    return _make_excel_response(dataset, filename)

# Exportación PDF para dueño
@app.route('/reportes/export/pdf', methods=['GET'], endpoint='exportar_reporte_pdf_dueno')
@login_required
@requiere_rol(2)
def exportar_reporte_pdf_dueno():
    reporte_tipo = request.args.get('reporte_tipo', 'animales_finca')
    finca_id = request.args.get('finca_id', type=int)
    desde = _parse_date(request.args.get('desde'))
    hasta = _parse_date(request.args.get('hasta'))
    dataset = _dataset_dueno(reporte_tipo, finca_id, desde, hasta)
    filename = f"{reporte_tipo}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}"
    return _make_pdf_response(dataset, filename)

@app.route('/admin/usuario/crear', methods=['GET', 'POST'])
@login_required
def crear_usuario_route():
    return crear_usuario()

@app.route('/admin/usuario/<int:usuario_id>/editar', methods=['GET', 'POST'])
@login_required
def editar_usuario_route(usuario_id):
    return editar_usuario(usuario_id)

@app.route('/admin/usuario/<int:usuario_id>/eliminar', methods=['POST'])
@login_required
def eliminar_usuario_route(usuario_id):
    return eliminar_usuario_controlador(usuario_id)

# Ruta para eliminar un usuario (solo accesible para el administrador root)
@app.route('/admin/eliminar-usuario/<int:usuario_id>', methods=['POST'])
@login_required
def eliminar_usuario(usuario_id):
    # Verificar que el usuario actual es administrador (tipo_usuario = 3)
    if current_user.tipo_usuario != 3:
        flash('No tienes permisos para realizar esta acción', 'danger')
        return redirect(url_for('dashboard'))
    
    usuario = Usuario.query.get_or_404(usuario_id)
    
    # No permitir eliminar al superadmin
    if usuario.tipo_usuario == 3:
        flash('No se puede eliminar al administrador del sistema', 'danger')
        return redirect(url_for('admin_usuarios'))
    
    try:
        db.session.delete(usuario)
        db.session.commit()
        flash(f'Usuario {usuario.nik_name} eliminado correctamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar el usuario: {str(e)}', 'danger')
    
    return redirect(url_for('admin_usuarios'))

# Ruta para que un usuario elimine su propia cuenta
@app.route('/mi-cuenta/eliminar', methods=['POST'])
@login_required
def eliminar_mi_cuenta():
    try:
        db.session.delete(current_user)
        db.session.commit()
        logout_user()
        flash('Tu cuenta ha sido eliminada correctamente', 'success')
        return redirect(url_for('login'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar tu cuenta: {str(e)}', 'danger')
        return redirect(url_for('mi_cuenta'))


# Rutas para gestión de animales
@app.route('/gestion_animales')
@login_required
def gestion_animales():
    return listar_animales()

# Nueva pestaña: Gestión de Producción
@app.route('/gestion_produccion', endpoint='gestion_produccion_route')
@login_required
def gestion_produccion_route():
    return gestion_produccion()

# Vista dedicada de producción por finca
@app.route('/gestion_produccion/finca/<int:finca_id>', endpoint='gestion_produccion_finca_route')
@login_required
def gestion_produccion_finca_route(finca_id):
    return gestion_produccion_finca(finca_id)

@app.route('/finca/<int:finca_id>/animales')
@login_required
def ver_animales_finca_route(finca_id):
    return ver_animales_finca(finca_id)

@app.route('/finca/<int:finca_id>/animales-fuera')
@login_required
def ver_animales_fuera_route(finca_id):
    return ver_animales_fuera(finca_id)

@app.route('/animales-fuera', endpoint='ver_animales_fuera_global_route')
@login_required
def ver_animales_fuera_global_route():
    return ver_animales_fuera_global()

@app.route('/animal/crear', methods=['GET', 'POST'])
@login_required
def crear_animal_route():
    return crear_animal()

@app.route('/animal/<int:animal_id>/editar', methods=['GET', 'POST'])
@login_required
def editar_animal_route(animal_id):
    return editar_animal(animal_id)

@app.route('/animal/<int:animal_id>/eliminar', methods=['POST'])
@login_required
def eliminar_animal_route(animal_id):
    return eliminar_animal(animal_id)

@app.route('/animal/<int:animal_id>', methods=['GET', 'POST'])
@login_required
def ver_animal_route(animal_id):
    return ver_animal(animal_id)

# Páginas dedicadas para peso, ciclo y salud
@app.route('/animal/<int:animal_id>/peso', methods=['GET', 'POST'], endpoint='animal_peso_route')
@login_required
def animal_peso_route(animal_id):
    return ver_peso_animal(animal_id)

@app.route('/animal/<int:animal_id>/ciclo', methods=['GET', 'POST'], endpoint='animal_ciclo_route')
@login_required
def animal_ciclo_route(animal_id):
    return ver_ciclo_animal(animal_id)

@app.route('/animal/<int:animal_id>/salud', methods=['GET', 'POST'], endpoint='animal_salud_route')
@login_required
def animal_salud_route(animal_id):
    return ver_salud_animal(animal_id)

# Página dedicada de gráficos
@app.route('/animal/<int:animal_id>/graficos', methods=['GET'], endpoint='animal_graficos_route')
@login_required
def animal_graficos_route(animal_id):
    return ver_graficos_animal(animal_id)

# Genealogía del animal
@app.route('/animal/<int:animal_id>/genealogia', endpoint='genealogia_animal_route')
@login_required
def genealogia_animal_route(animal_id):
    return genealogia_animal(animal_id)

# Procedimientos del animal (salud y sexuales)
@app.route('/animal/<int:animal_id>/procedimientos', methods=['GET', 'POST'], endpoint='procedimientos_animal_route')
@login_required
def procedimientos_animal_route(animal_id):
    return procedimientos_animal(animal_id)

# Consumo del animal (leche y consumibles)
@app.route('/animal/<int:animal_id>/consumo', methods=['GET', 'POST'], endpoint='animal_consumo_route')
@login_required
def animal_consumo_route(animal_id):
    return consumo_animal(animal_id)

# Biológicos del animal (sexual, semen, embriones)
@app.route('/animal/<int:animal_id>/biologicos', methods=['GET', 'POST'], endpoint='animal_biologicos_route')
@login_required
def animal_biologicos_route(animal_id):
    return biologicos_animal(animal_id)

# Producción registrada del animal
@app.route('/animal/<int:animal_id>/produccion', methods=['GET', 'POST'], endpoint='animal_produccion_route')
@login_required
def animal_produccion_route(animal_id):
    return ver_produccion_animal(animal_id)

# Historial de procedimientos del animal
@app.route('/animal/<int:animal_id>/historial', endpoint='historial_procedimientos_route')
@login_required
def historial_procedimientos_route(animal_id):
    return historial_procedimientos(animal_id)

# Documentos genéticos
@app.route('/animal/<int:animal_id>/documentos_geneticos', endpoint='documentos_geneticos_route')
@login_required
def documentos_geneticos_route(animal_id):
    return documentos_geneticos(animal_id)

@app.route('/animal/<int:animal_id>/documentos_geneticos/agregar', methods=['GET', 'POST'], endpoint='agregar_documento_genetico_route')
@login_required
def agregar_documento_genetico_route(animal_id):
    return agregar_documento_genetico(animal_id)

@app.route('/documento_genetico/<int:documento_id>', endpoint='ver_documento_genetico_route')
@login_required
def ver_documento_genetico_route(documento_id):
    return ver_documento_genetico(documento_id)

# Descargar documento genético
@app.route('/documento_genetico/<int:documento_id>/descargar', endpoint='descargar_documento_genetico_route')
@login_required
def descargar_documento_genetico_route(documento_id):
    return descargar_documento_genetico(documento_id)

# Eliminar documento genético
@app.route('/documento_genetico/<int:documento_id>/eliminar', methods=['POST'], endpoint='eliminar_documento_genetico_route')
@login_required
def eliminar_documento_genetico_route(documento_id):
    return eliminar_documento_genetico(documento_id)

# Eliminar procedimientos
@app.route('/animal/<int:animal_id>/servicio_salud/<int:servicio_id>/eliminar', methods=['POST'], endpoint='eliminar_servicio_salud_route')
@login_required
def eliminar_servicio_salud_route(animal_id, servicio_id):
    return eliminar_servicio_salud(animal_id, servicio_id)

@app.route('/animal/<int:animal_id>/servicio_sexual/<int:servicio_id>/eliminar', methods=['POST'], endpoint='eliminar_servicio_sexual_route')
@login_required
def eliminar_servicio_sexual_route(animal_id, servicio_id):
    return eliminar_servicio_sexual(animal_id, servicio_id)

@app.route('/animal/<int:animal_id>/peso/<int:registro_id>/eliminar', methods=['POST'], endpoint='eliminar_registro_peso_route')
@login_required
def eliminar_registro_peso_route(animal_id, registro_id):
    return eliminar_registro_peso(animal_id, registro_id)

# Ruta para servir la foto del animal desde BD
@app.route('/animal/<int:animal_id>/foto')
@login_required
def ver_foto_animal_route(animal_id):
    return ver_foto_animal(animal_id)

# Ruta para servir la foto del usuario
@app.route('/usuario/<int:usuario_id>/foto')
@login_required
def ver_foto_usuario_route(usuario_id):
    return ver_foto_usuario(usuario_id)

# API endpoints para animales
@app.route('/api/animales/finca/<int:finca_id>')
@login_required
def api_animales_por_finca(finca_id):
    return obtener_animales_por_finca(finca_id)

@app.route('/api/animales/sexo/<sexo>')
@login_required
def api_animales_por_sexo(sexo):
    return obtener_animales_por_sexo(sexo)

@app.route('/api/potreros-por-finca')
@login_required
def api_potreros_por_finca():
    return get_potreros_por_finca()

@app.route('/api/potrero/default-tipo-uso')
@login_required
def api_default_tipo_uso_potrero_route():
    return api_default_tipo_uso_potrero()

# API para grupos activos por potrero
@app.route('/api/grupos-activos-por-potrero')
@login_required
def api_grupos_activos_por_potrero_route():
    return api_grupos_activos_por_potrero()

# API para grupos por finca (fallback)
@app.route('/api/grupos-por-finca')
@login_required
def api_grupos_por_finca_route():
    return api_grupos_por_finca()

# API para madurez sexual por raza
@app.route('/api/raza/<int:raza_id>/madurez-sexual')
@login_required
def api_madurez_sexual_por_raza_route(raza_id):
    return api_madurez_sexual_por_raza(raza_id)

# API para verificar nombre de animal duplicado
@app.route('/api/animal/existe-nombre')
@login_required
def api_existe_nombre_animal_route():
    return api_existe_nombre_animal()

# Endpoints adicionales para gestión de animales en potreros
@app.route('/api/animales-disponibles')
@login_required
def api_animales_disponibles():
    return get_animales_disponibles()

@app.route('/api/animales-potrero')
@login_required
def api_animales_potrero():
    return get_animales_potrero()

@app.route('/api/asignar-animales-potrero', methods=['POST'])
@login_required
def api_asignar_animales_potrero():
    return asignar_animales_potrero()

# Aplicar decoradores de rol a las rutas de gestión de animales
admin_usuarios = requiere_rol(3)(admin_usuarios)
crear_usuario_route = requiere_rol(3)(crear_usuario_route)
editar_usuario_route = requiere_rol(3)(editar_usuario_route)
eliminar_usuario_route = requiere_rol(3)(eliminar_usuario_route)
admin_reportes = requiere_rol(3)(admin_reportes)

# Ruta para administración de fincas (solo para admin)
@app.route('/admin/fincas')
@login_required
@requiere_rol(3)
def admin_fincas():
    from modelo.models import Finca, Usuario, UsuarioFinca
    
    # Obtener todas las fincas con información del propietario a través de UsuarioFinca
    fincas = db.session.query(Finca, Usuario).join(
        UsuarioFinca, Finca.id_finca == UsuarioFinca.finca_id
    ).join(
        Usuario, UsuarioFinca.usuario_id == Usuario.id
    ).all()
    
    # Obtener todos los usuarios dueños de finca (tipo_usuario = 2)
    usuarios_duenos = Usuario.query.filter(Usuario.tipo_usuario == 2).all()
    
    # Estadísticas
    total_fincas = Finca.query.count()
    total_propietarios = db.session.query(Usuario.id).filter(Usuario.tipo_usuario == 2).count()
    
    return render_template('root/admin_fincas.html', 
                         fincas=fincas,
                         usuarios_duenos=usuarios_duenos,
                         total_fincas=total_fincas,
                         fincas_activas=total_fincas,  # Usar total_fincas ya que no hay campo estado
                         total_propietarios=total_propietarios)

# Rutas para gestión de fincas
@app.route('/crear-finca', methods=['GET', 'POST'])
@login_required
def crear_finca_route():
    return crear_finca()

@app.route('/mis-fincas')
@login_required
def mis_fincas():
    return listar_fincas()

@app.route('/finca/<int:finca_id>/editar', methods=['GET', 'POST'])
@login_required
def editar_finca_route(finca_id):
    return editar_finca(finca_id)

@app.route('/finca/<int:finca_id>/eliminar', methods=['POST'])
@login_required
def eliminar_finca_route(finca_id):
    return eliminar_finca(finca_id)

@app.route('/finca/gestionar/<int:finca_id>')
@login_required
def gestionar_finca_route(finca_id):
    return gestionar_finca(finca_id)

@app.route('/obtener-fincas-usuario')
@login_required
def obtener_fincas_usuario_route():
    return obtener_fincas_usuario()

@app.route('/finca/<int:finca_id>')
@login_required
def ver_finca_route(finca_id):
    return ver_finca(finca_id)

# Rutas de gestión de potreros
@app.route('/finca/<int:finca_id>/potrero/crear', methods=['GET', 'POST'])
@login_required
def crear_potrero_route(finca_id):
    return crear_potrero(finca_id)

@app.route('/potrero/<int:potrero_id>/editar', methods=['GET', 'POST'])
@login_required
def editar_potrero_route(potrero_id):
    return editar_potrero(potrero_id)

@app.route('/potrero/<int:potrero_id>/eliminar', methods=['POST'])
@login_required
def eliminar_potrero_route(potrero_id):
    return eliminar_potrero(potrero_id)

# Nuevas páginas de potrero
@app.route('/potrero/<int:potrero_id>/ver')
@login_required
def ver_potrero_route(potrero_id):
    return ver_potrero(potrero_id)

@app.route('/potrero/<int:potrero_id>/agregar-animales')
@login_required
def agregar_animales_potrero_route(potrero_id):
    return agregar_animales_potrero(potrero_id)

# Endpoint para guardar rotación desde el modal
@app.route('/guardar-rotacion', methods=['POST'])
@login_required
def guardar_rotacion_route():
    return guardar_rotacion()

# Gestión de grupos de animales
@app.route('/finca/<int:finca_id>/grupos', methods=['GET', 'POST'])
@login_required
def listar_grupos_finca_route(finca_id):
    return listar_grupos_finca(finca_id)

@app.route('/grupo/<int:grupo_id>/gestionar')
@login_required
def gestionar_grupo_route(grupo_id):
    return gestionar_grupo(grupo_id)

@app.route('/grupo/<int:grupo_id>/eliminar', methods=['POST'], endpoint='eliminar_grupo_route')
@login_required
def eliminar_grupo_route(grupo_id):
    return eliminar_grupo(grupo_id)

@app.route('/api/grupo/<int:grupo_id>/agregar-animal', methods=['POST'])
@login_required
def api_agregar_animal_a_grupo_route(grupo_id):
    return api_agregar_animal_a_grupo(grupo_id)

@app.route('/api/grupo/<int:grupo_id>/quitar-animal', methods=['POST'])
@login_required
def api_quitar_animal_de_grupo_route(grupo_id):
    return api_quitar_animal_de_grupo(grupo_id)

# Aplicar decoradores de rol a las rutas de gestión de animales
gestion_animales = requiere_rol(2)(gestion_animales)
gestion_produccion_route = requiere_rol(2)(gestion_produccion_route)
gestion_produccion_finca_route = requiere_rol(2)(gestion_produccion_finca_route)
crear_animal_route = requiere_rol(2)(crear_animal_route)
editar_animal_route = requiere_rol(2)(editar_animal_route)
eliminar_animal_route = requiere_rol(2)(eliminar_animal_route)
ver_animal_route = requiere_rol(2)(ver_animal_route)
genealogia_animal_route = requiere_rol(2)(genealogia_animal_route)
procedimientos_animal_route = requiere_rol(2)(procedimientos_animal_route)
historial_procedimientos_route = requiere_rol(2)(historial_procedimientos_route)
ver_foto_animal_route = requiere_rol(2)(ver_foto_animal_route)
documentos_geneticos_route = requiere_rol(2)(documentos_geneticos_route)
agregar_documento_genetico_route = requiere_rol(2)(agregar_documento_genetico_route)
ver_documento_genetico_route = requiere_rol(2)(ver_documento_genetico_route)
descargar_documento_genetico_route = requiere_rol(2)(descargar_documento_genetico_route)
eliminar_documento_genetico_route = requiere_rol(2)(eliminar_documento_genetico_route)
eliminar_servicio_salud_route = requiere_rol(2)(eliminar_servicio_salud_route)
eliminar_servicio_sexual_route = requiere_rol(2)(eliminar_servicio_sexual_route)
eliminar_registro_peso_route = requiere_rol(2)(eliminar_registro_peso_route)
animal_consumo_route = requiere_rol(2)(animal_consumo_route)
animal_biologicos_route = requiere_rol(2)(animal_biologicos_route)
animal_produccion_route = requiere_rol(2)(animal_produccion_route)
animal_peso_route = requiere_rol(2)(animal_peso_route)
animal_ciclo_route = requiere_rol(2)(animal_ciclo_route)
animal_salud_route = requiere_rol(2)(animal_salud_route)
animal_graficos_route = requiere_rol(2)(animal_graficos_route)

# Rutas de gestión de trabajadores (Dueño)
@app.route('/trabajadores')
@login_required
def gestionar_trabajadores_route():
    return listar_trabajadores_dueno()

@app.route('/trabajadores/crear', methods=['GET', 'POST'])
@login_required
def crear_trabajador_dueno_route():
    return crear_trabajador_dueno()

@app.route('/trabajadores/actualizar-permisos', methods=['POST'])
@login_required
def actualizar_permiso_trabajador_route():
    return actualizar_permiso_trabajador()

# Página dedicada de trabajadores por finca
@app.route('/finca/<int:finca_id>/trabajadores')
@login_required
@requiere_rol(2)
def gestionar_trabajadores_finca_route(finca_id):
    return gestionar_trabajadores_finca(finca_id)

# Ruta y vista de administración de trabajador
@app.route('/trabajador/<int:usuario_id>/administrar', methods=['GET', 'POST'])
@login_required
@requiere_rol(2)
def administrar_trabajador_route(usuario_id):
    return ver_trabajador_admin(usuario_id) if request.method == 'GET' else actualizar_trabajador_admin(usuario_id)

# Eliminar trabajador
@app.route('/trabajador/<int:usuario_id>/eliminar', methods=['POST'], endpoint='eliminar_trabajador_route')
@login_required
@requiere_rol(2)
def eliminar_trabajador_route(usuario_id):
    return eliminar_trabajador(usuario_id)

# Aplicar decoradores de rol a rutas de trabajadores
gestionar_trabajadores_route = requiere_rol(2)(gestionar_trabajadores_route)
crear_trabajador_dueno_route = requiere_rol(2)(crear_trabajador_dueno_route)
actualizar_permiso_trabajador_route = requiere_rol(2)(actualizar_permiso_trabajador_route)
ver_foto_usuario_route = requiere_rol(1)(ver_foto_usuario_route)

# Asignar/Quitar trabajador en finca
@app.route('/finca/<int:finca_id>/trabajador/<int:usuario_id>/asignar', methods=['POST'])
@login_required
@requiere_rol(2)
def asignar_trabajador_finca_route(finca_id, usuario_id):
    return asignar_trabajador_finca(finca_id, usuario_id)

@app.route('/finca/<int:finca_id>/trabajador/<int:usuario_id>/quitar', methods=['POST'])
@login_required
@requiere_rol(2)
def quitar_trabajador_finca_route(finca_id, usuario_id):
    return quitar_trabajador_finca(finca_id, usuario_id)

# Permisos por funcionalidad (por finca y trabajador)
@app.route('/finca/<int:finca_id>/trabajador/<int:usuario_id>/permisos', methods=['GET'])
@login_required
@requiere_rol(2)
def obtener_permisos_finca_trabajador_route(finca_id, usuario_id):
    return obtener_permisos_finca_trabajador(finca_id, usuario_id)

@app.route('/finca/<int:finca_id>/trabajador/permisos/actualizar', methods=['POST'])
@login_required
@requiere_rol(2)
def actualizar_permiso_finca_trabajador_route(finca_id):
    return actualizar_permiso_finca_trabajador()

# Editar documento genético
@app.route('/documento_genetico/<int:documento_id>/editar', methods=['GET', 'POST'], endpoint='editar_documento_genetico_route')
@login_required
def editar_documento_genetico_route(documento_id):
    return editar_documento_genetico(documento_id)

editar_documento_genetico_route = requiere_rol(2)(editar_documento_genetico_route)
ver_animales_finca_route = requiere_rol(2)(ver_animales_finca_route)
ver_animales_fuera_route = requiere_rol(2)(ver_animales_fuera_route)
ver_animales_fuera_global_route = requiere_rol(2)(ver_animales_fuera_global_route)
api_animales_por_finca = requiere_rol(2)(api_animales_por_finca)
api_animales_por_sexo = requiere_rol(2)(api_animales_por_sexo)
api_animales_disponibles = requiere_rol(2)(api_animales_disponibles)
api_animales_potrero = requiere_rol(2)(api_animales_potrero)
api_asignar_animales_potrero = requiere_rol(2)(api_asignar_animales_potrero)
api_default_tipo_uso_potrero_route = requiere_rol(2)(api_default_tipo_uso_potrero_route)
api_madurez_sexual_por_raza_route = requiere_rol(2)(api_madurez_sexual_por_raza_route)
api_existe_nombre_animal_route = requiere_rol(2)(api_existe_nombre_animal_route)

# Aplicar decoradores de rol a las rutas de gestión de fincas
crear_finca_route = requiere_rol(2)(crear_finca_route)
mis_fincas = requiere_rol(2)(mis_fincas)
editar_finca_route = requiere_rol(2)(editar_finca_route)
eliminar_finca_route = requiere_rol(2)(eliminar_finca_route)
gestionar_finca_route = requiere_rol(2)(gestionar_finca_route)
obtener_fincas_usuario_route = requiere_rol(2)(obtener_fincas_usuario_route)
ver_finca_route = requiere_rol(2)(ver_finca_route)

# Aplicar decoradores de rol a las rutas de gestión de potreros
crear_potrero_route = requiere_rol(2)(crear_potrero_route)
editar_potrero_route = requiere_rol(2)(editar_potrero_route)
eliminar_potrero_route = requiere_rol(2)(eliminar_potrero_route)
ver_potrero_route = requiere_rol(2)(ver_potrero_route)
agregar_animales_potrero_route = requiere_rol(2)(agregar_animales_potrero_route)
guardar_rotacion_route = requiere_rol(2)(guardar_rotacion_route)
listar_grupos_finca_route = requiere_rol(2)(listar_grupos_finca_route)
gestionar_grupo_route = requiere_rol(2)(gestionar_grupo_route)
api_agregar_animal_a_grupo_route = requiere_rol(2)(api_agregar_animal_a_grupo_route)
api_quitar_animal_de_grupo_route = requiere_rol(2)(api_quitar_animal_de_grupo_route)
api_grupos_activos_por_potrero_route = requiere_rol(2)(api_grupos_activos_por_potrero_route)
api_grupos_por_finca_route = requiere_rol(2)(api_grupos_por_finca_route)
eliminar_grupo_route = requiere_rol(2)(eliminar_grupo_route)

# Eliminada: Ruta /guardar-potrero que procesaba el formulario modal mediante AJAX
if __name__ == '__main__':
    app.run(debug=True, host="0.0.0.0", port=5000)
    
    

